# -*- coding: utf-8 -*-
"""
Ensambla index.html a partir de app.src.html:
  /*ALUMNOS*/   -> arreglo JS con los alumnos, leido del Excel de _datos\
  /*FONTS*/     -> @font-face de Iberoamericana (woff2 en base64)
  <!--LOGO-->   -> SVG del logo oficial IBERO Tijuana, inline, color heredado
  /*LOGO_SVG*/  -> el mismo SVG como cadena JS (blanco) para el canvas

Uso:  python build.py
"""
import os, re, json, base64, io, glob

AQUI = os.path.dirname(os.path.abspath(__file__))
PROY = os.path.dirname(AQUI)
RECURSOS = os.environ.get(
    "IBERO_RECURSOS",
    r"C:\Users\leonardo.gonzalez\Documents\IBERO - DG\2026\00_RECURSOS")
SVG_ORIG = os.path.join(RECURSOS, "Logos", "PROMOCIONALES", "IBEROTijuana_LTR.svg")
FUENTES = os.path.join(RECURSOS, "Tipograf\u00edas", "Iberoamericana")

# Copia de lo ya generado (fuentes en base64 y logotipo), para poder compilar
# en una PC que no tenga la carpeta RECURSOS. Se regenera sola cuando si esta.
CACHE = os.path.join(AQUI, "cache")
HAY_RECURSOS = os.path.isdir(FUENTES) and os.path.isfile(SVG_ORIG)


def cache_lee(nombre):
    p = os.path.join(CACHE, nombre)
    if not os.path.isfile(p):
        raise SystemExit(
            "ERROR: no esta RECURSOS (%s) ni el respaldo _build\\cache\\%s.\n"
            "       Define IBERO_RECURSOS con la ruta correcta, o recupera el cache."
            % (RECURSOS, nombre))
    return open(p, encoding="utf-8").read()


def cache_escribe(nombre, texto):
    os.makedirs(CACHE, exist_ok=True)
    open(os.path.join(CACHE, nombre), "w", encoding="utf-8").write(texto)

VIEWBOX = "48 62 680 270"          # recorte al logotipo, sin el fondo rojo
ASPECTO = (680, 270)

# ── 1. FUENTES ───────────────────────────────────────────────────────────
UNI = ("U+0020-007F,U+00A0-00FF,U+2013-2014,U+2018-201D,"
       "U+2022,U+2026,U+00D7,U+2713,U+2192")
PESOS = [("Regular", 400), ("Medium", 500), ("Bold", 700), ("Black", 900)]


def unicodes():
    out = []
    for r in UNI.split(","):
        r = r[2:]
        if "-" in r:
            a, b = r.split("-")
            out += list(range(int(a, 16), int(b, 16) + 1))
        else:
            out.append(int(r, 16))
    return out


def css_fuentes():
    if not HAY_RECURSOS:
        print("  sin RECURSOS: uso las fuentes del cache")
        return cache_lee("fuentes.css")
    from fontTools.ttLib import TTFont
    from fontTools.subset import Subsetter, Options
    bloques = []
    for nombre, peso in PESOS:
        f = TTFont(os.path.join(FUENTES, "Iberoamericana-%s.ttf" % nombre))
        o = Options()
        o.layout_features = ["kern", "liga", "calt", "ccmp", "locl"]
        o.notdef_outline = True
        o.name_IDs = ["*"]
        o.name_legacy = True
        o.drop_tables += ["DSIG"]
        s = Subsetter(options=o)
        s.populate(unicodes=unicodes())
        s.subset(f)
        f.flavor = "woff2"
        buf = io.BytesIO()
        f.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()
        print("  fuente %-8s %6.1f KB" % (nombre, len(buf.getvalue()) / 1024))
        bloques.append(
            "@font-face{font-family:'Iberoamericana';font-style:normal;"
            "font-weight:%d;font-display:swap;"
            "src:url(data:font/woff2;base64,%s) format('woff2')}" % (peso, b64))
    css = "\n".join(bloques)
    cache_escribe("fuentes.css", css)
    return css


# ── 2. LOGO ──────────────────────────────────────────────────────────────
def logo_svg(color):
    """Devuelve el logotipo limpio: sin el rectangulo rojo ni la nota Pantone."""
    src = open(SVG_ORIG, encoding="utf-8").read()
    cuerpo = re.search(r"<g>.*</g>", src, re.S).group(0)      # solo los grupos
    cuerpo = cuerpo.replace('class="st1"', 'fill="%s"' % color)
    cuerpo = re.sub(r'\s*class="st0"', "", cuerpo)
    return ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="%s" '
            'width="%d" height="%d" role="img" aria-label="IBERO Tijuana">%s</svg>'
            % (VIEWBOX, ASPECTO[0], ASPECTO[1], cuerpo))


def logos():
    """(svg inline con class="logo", cadena JS del mismo logo en blanco)."""
    if not HAY_RECURSOS:
        print("  sin RECURSOS: uso el logotipo del cache")
        return cache_lee("logo_inline.svg"), cache_lee("logo_js.txt")
    inline = logo_svg("currentColor").replace('<svg ', '<svg class="logo" ')
    js = json.dumps(logo_svg("#FFFFFF"))
    cache_escribe("logo_inline.svg", inline)
    cache_escribe("logo_js.txt", js)
    return inline, js


# ── 3. ALUMNOS ───────────────────────────────────────────────────────────
# Encabezados esperados en la hoja (se buscan por nombre, no por posicion).
COLS = {
    "carrera": ["carrera"],
    "cuenta":  ["cuenta", "matricula", "matrícula"],
    "ap":      ["apellido paterno"],
    "am":      ["apellido materno", "apellido maerno"],   # el Excel trae la errata
    "nom":     ["nombre", "nombre(s)", "nombres"],
    "grupo":   ["grupo"],
    "salon":   ["salon", "salón", "aula"],
}


def limpia(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def leer_alumnos():
    import openpyxl
    libros = glob.glob(os.path.join(PROY, "_datos", "*.xlsx"))
    if not libros:
        raise SystemExit("ERROR: no hay ningun .xlsx en _datos\\")
    if len(libros) > 1:
        raise SystemExit("ERROR: hay %d archivos en _datos\\; deja solo uno." % len(libros))
    ws = openpyxl.load_workbook(libros[0], data_only=True).worksheets[0]
    print("Leyendo", os.path.basename(libros[0]), "hoja", repr(ws.title))

    # localizar la fila de encabezados
    idx, cab = None, {}
    for n, fila in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        celdas = {limpia(c).lower(): i for i, c in enumerate(fila) if limpia(c)}
        hall = {k: next((celdas[a] for a in alias if a in celdas), None)
                for k, alias in COLS.items()}
        if hall["nom"] is not None and hall["grupo"] is not None:
            idx, cab = n, hall
            break
    if idx is None:
        raise SystemExit("ERROR: no encontre los encabezados (Nombre / Grupo).")
    faltan = [k for k, v in cab.items() if v is None]
    if faltan:
        print("  aviso: sin columna para", faltan)

    def val(fila, k):
        i = cab.get(k)
        return limpia(fila[i]) if i is not None and i < len(fila) else ""

    alumnos = []
    for fila in ws.iter_rows(min_row=idx + 1, values_only=True):
        nom, ap, am = val(fila, "nom"), val(fila, "ap"), val(fila, "am")
        if not (nom or ap):
            continue
        alumnos.append({
            "nombre":  " ".join(p for p in (nom, ap, am) if p),
            "cuenta":  val(fila, "cuenta"),
            "carrera": val(fila, "carrera"),
            "grupo":   val(fila, "grupo"),
            "salon":   val(fila, "salon"),
        })

    alumnos.sort(key=lambda a: a["nombre"])
    for campo, etiqueta in (("grupo", "grupo"), ("salon", "salon")):
        sin = [a for a in alumnos if not a[campo]]
        if sin:
            print("  AVISO: %d sin %s ->" % (len(sin), etiqueta),
                  [a["nombre"] for a in sin][:5])
    print("  %d alumnos · %d carreras · grupos %s"
          % (len(alumnos),
             len({a["carrera"] for a in alumnos}),
             ", ".join(sorted({a["grupo"] for a in alumnos if a["grupo"]}))))

    # un grupo deberia ocupar un solo salon; avisar si no es asi
    for g in sorted({a["grupo"] for a in alumnos if a["grupo"]}):
        salones = sorted({a["salon"] for a in alumnos if a["grupo"] == g and a["salon"]})
        if len(salones) > 1:
            print("  AVISO: el grupo %s aparece en varios salones -> %s"
                  % (g, ", ".join(salones)))

    lineas = ["{nombre:%s, cuenta:%s, carrera:%s, grupo:%s, salon:%s}" % (
        json.dumps(a["nombre"], ensure_ascii=False),
        json.dumps(a["cuenta"], ensure_ascii=False),
        json.dumps(a["carrera"], ensure_ascii=False),
        json.dumps(a["grupo"], ensure_ascii=False),
        json.dumps(a["salon"], ensure_ascii=False)) for a in alumnos]
    return "[\n " + ",\n ".join(lineas) + "\n]"


# ── 4. ENSAMBLE ──────────────────────────────────────────────────────────
def main():
    src = open(os.path.join(AQUI, "app.src.html"), encoding="utf-8").read()

    src = src.replace("/*ALUMNOS*/[]", leer_alumnos())

    print("Generando fuentes...")
    src = src.replace("/*FONTS*/", css_fuentes())

    inline, logo_js = logos()
    src = src.replace("<!--LOGO-->", inline)
    src = src.replace('/*LOGO_SVG*/""', logo_js)

    for falta in ("/*ALUMNOS*/", "/*FONTS*/", "<!--LOGO-->", "/*LOGO_SVG*/"):
        if falta in src:
            raise SystemExit("ERROR: quedo sin sustituir %s" % falta)

    dest = os.path.join(PROY, "index.html")
    open(dest, "w", encoding="utf-8").write(src)
    print("OK ->", dest, "%.1f KB" % (os.path.getsize(dest) / 1024))


if __name__ == "__main__":
    main()
